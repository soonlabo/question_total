import glob
import pandas as pd
import openpyxl

from openpyxl import load_workbook

#読み込むEXCELファイル
inputFiles=glob.glob('C:\\Users\\user\\Desktop\\アンケート\\input\\*.xlsx')

#書き込むEXCELファイル
outputFile=openpyxl.load_workbook('C:\\Users\\user\\Desktop\\アンケート\\集計表.xlsx')
syuukei_sheet=outputFile['集計']
iken_sheet=outputFile['意見']

print(inputFiles)

for count, inputFile in enumerate(inputFiles, 1):
    input_workbook = load_workbook(filename=inputFile, read_only=True)
    input_sheet = input_workbook['Sheet1']

    # 仕事内容
    name = input_sheet['F10'].value
    gaikinsaki = input_sheet['U10'].value
    sagyounaiyou = input_sheet['F12'].value
    siyougijyutu = input_sheet['F14'].value
    koyoukeitai = input_sheet['F16'].value
    CDA = input_sheet['F18'].value
    kinnzokunennsuu = input_sheet['F20'].value
    tantou = input_sheet['F22'].value

    # 外勤先
    A1 = input_sheet['Z27'].value
    B1 = input_sheet['Z28'].value
    C1 = input_sheet['Z29'].value
    D1 = input_sheet['Z30'].value
    E1 = input_sheet['Z31'].value


    A2 = input_sheet['Z36'].value
    B2 = input_sheet['Z37'].value
    C2 = input_sheet['Z38'].value
    D2 = input_sheet['Z39'].value
    E2 = input_sheet['Z40'].value
    F2 = input_sheet['Z41'].value
    G2 = input_sheet['Z42'].value

    # 意見
    iken = input_sheet['A46'].value

    #書き込み
    syuukei_sheet.cell(row=count, column=1).value = name
    syuukei_sheet.cell(row=count, column=2).value = A1
    syuukei_sheet.cell(row=count, column=3).value = B1
    syuukei_sheet.cell(row=count, column=4).value = C1
    syuukei_sheet.cell(row=count, column=5).value = D1
    syuukei_sheet.cell(row=count, column=6).value = E1
    syuukei_sheet.cell(row=count, column=7).value = A2
    syuukei_sheet.cell(row=count, column=8).value = B2
    syuukei_sheet.cell(row=count, column=9).value = C2
    syuukei_sheet.cell(row=count, column=10).value = D2
    syuukei_sheet.cell(row=count, column=11).value = E2
    syuukei_sheet.cell(row=count, column=12).value = F2
    syuukei_sheet.cell(row=count, column=13).value = G2
    iken_sheet.cell(row=count, column=1).value = name
    iken_sheet.cell(row=count, column=2).value = iken

    # 保存する
    outputFile.save('C:\\Users\\user\\Desktop\\アンケート\\集計表.xlsx')

    input_workbook.close()
